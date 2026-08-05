# -*- coding: utf-8 -*-
"""
main.py
-------
Application de traitement de la paie mensuelle -- Burkina Faso.

Deux niveaux d'accès :
  - Administrateur : mot de passe fixe (modifiable), accès aux paramètres
    de paie et au mot de passe Utilisateur.
  - Utilisateur : mot de passe qui change automatiquement chaque mois,
    accès à la saisie des employés et au calcul de la paie.

Lancer avec :  python main.py
"""

import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from dataclasses import replace

import auth
import storage
from payroll_engine import Employee, compute_payslip, find_base_for_target_net, DEFAULT_PARAMS

APP_TITLE = "Paie Burkina — Traitement des salaires mensuels"
PAID_SOFTWARE_NOTICE = "Ce logiciel de paie est payant : consultanter280@gmail.com"

MOIS_FR = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
           "Août", "Septembre", "Octobre", "Novembre", "Décembre"]


def current_period_key():
    today = datetime.date.today()
    return f"{today.year:04d}-{today.month:02d}"


def normalize_period(text, default=None):
    """Accepte 'MM/AAAA', 'AAAA-MM', 'MM-AAAA', 'AAAA/MM', ou une date Excel,
    et renvoie 'AAAA-MM'. Renvoie `default` (ou la période en cours) si le
    texte est vide/invalide."""
    default = default or current_period_key()
    if isinstance(text, (datetime.date, datetime.datetime)):
        return f"{text.year:04d}-{text.month:02d}"
    text = (text or "").strip()
    if not text:
        return default
    parts = text.replace("/", "-").split("-")
    if len(parts) != 2:
        return default
    a, b = parts
    try:
        if len(a) == 4:
            year, month = int(a), int(b)
        else:
            month, year = int(a), int(b)
        if not (1 <= month <= 12):
            return default
        return f"{year:04d}-{month:02d}"
    except ValueError:
        return default


def format_period(period_key):
    try:
        year, month = period_key.split("-")
        return f"{MOIS_FR[int(month) - 1]} {year}"
    except Exception:
        return period_key or ""


def fmt_amount(v):
    """Formate un montant avec espace comme séparateur de milliers et sans
    décimales inutiles (150000.0 -> '150 000'). Les valeurs non numériques
    (texte, None...) sont renvoyées telles quelles."""
    if v is None or v == "":
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    return f"{f:,.0f}".replace(",", " ")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1250x720")
        self.minsize(1000, 620)
        try:
            self.state("zoomed")  # démarre en fenêtre maximisée sous Windows
        except tk.TclError:
            pass

        # IMPORTANT : sans ceci, toute erreur inattendue survenant dans un
        # bouton/callback Tkinter est silencieuse (surtout dans un .exe
        # compilé en mode "fenêtre", sans console) -- l'utilisateur voit
        # juste "rien ne se passe". On l'affiche désormais dans une boîte
        # de dialogue explicite, avec le détail technique, pour pouvoir
        # diagnostiquer immédiatement.
        self.report_callback_exception = self._show_error

        try:
            self.config_data = storage.load()
        except Exception as exc:
            messagebox.showerror(
                "Erreur au démarrage",
                "Impossible de charger/créer le fichier de données local.\n\n"
                f"Détail technique : {exc}\n\n"
                "Vérifiez que l'application a le droit d'écrire dans votre "
                "dossier utilisateur (essayez de la lancer en tant "
                "qu'administrateur, ou vérifiez qu'un antivirus ne la bloque pas).")
            raise
        self.role = None  # "admin" ou "user"

        self.container = ttk.Frame(self)
        self.container.pack(fill="both", expand=True)

        self.show_login()

    def _show_error(self, exc_type, exc_value, exc_tb):
        import traceback
        detail = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        messagebox.showerror(
            "Erreur inattendue",
            f"Une erreur s'est produite :\n\n{exc_value}\n\n"
            "Détail technique (à transmettre au support si le problème persiste) :\n"
            f"{detail[-1200:]}")

    # ------------------------------------------------------------------
    def clear(self):
        for w in self.container.winfo_children():
            w.destroy()

    def show_login(self):
        self.role = None
        self.clear()
        LoginScreen(self.container, self)

    def show_main(self):
        self.clear()
        MainScreen(self.container, self)


# ==========================================================================
# ÉCRAN DE CONNEXION
# ==========================================================================

class LoginScreen(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self.pack(fill="both", expand=True)

        banner = tk.Label(self, text=PAID_SOFTWARE_NOTICE, fg="white", bg="#b8860b",
                           font=("Segoe UI", 10, "bold"), pady=8)
        banner.pack(fill="x", side="top")

        center = ttk.Frame(self)
        center.pack(expand=True)

        ttk.Label(center, text="Paie Burkina", font=("Segoe UI", 22, "bold")).pack(pady=(40, 4))
        ttk.Label(center, text="Traitement des salaires mensuels",
                  font=("Segoe UI", 11)).pack(pady=(0, 24))

        form = ttk.Frame(center)
        form.pack()

        ttk.Label(form, text="Rôle :").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.role_var = tk.StringVar(value="Utilisateur")
        role_combo = ttk.Combobox(form, textvariable=self.role_var,
                                   values=["Utilisateur", "Administrateur"],
                                   state="readonly", width=20)
        role_combo.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        ttk.Label(form, text="Mot de passe :").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.pwd_var = tk.StringVar()
        self.show_pwd_var = tk.BooleanVar(value=False)
        pwd_entry = ttk.Entry(form, textvariable=self.pwd_var, show="•", width=23)
        pwd_entry.grid(row=1, column=1, sticky="w", padx=6, pady=6)
        pwd_entry.bind("<Return>", lambda e: self.try_login())
        pwd_entry.focus_set()

        def toggle_show():
            pwd_entry.config(show="" if self.show_pwd_var.get() else "•")

        ttk.Checkbutton(form, text="Afficher le mot de passe", variable=self.show_pwd_var,
                         command=toggle_show).grid(row=2, column=1, sticky="w", padx=6)

        ttk.Button(center, text="Se connecter", command=self.try_login).pack(pady=16)

        info = ttk.Label(
            center,
            text="Le mot de passe Utilisateur change automatiquement chaque\n"
                 "mois. Contactez l'administrateur pour l'obtenir.\n"
                 "Clavier AZERTY : pensez à Maj (Shift) pour taper les chiffres.",
            justify="center", foreground="#555")
        info.pack(pady=(6, 0))

        footer = tk.Label(self, text=PAID_SOFTWARE_NOTICE, fg="#b8860b",
                           font=("Segoe UI", 8, "italic"))
        footer.pack(side="bottom", pady=6)

    def try_login(self):
        role = self.role_var.get()
        pwd = self.pwd_var.get().strip()
        cfg = self.app.config_data

        if role == "Administrateur":
            ok = auth.verify_password(pwd, cfg["admin_salt"], cfg["admin_hash"])
            if ok:
                self.app.role = "admin"
                self.app.show_main()
            else:
                messagebox.showerror("Connexion refusée",
                                      "Mot de passe administrateur incorrect.\n\n"
                                      "Astuce : cochez « Afficher le mot de passe » pour vérifier "
                                      "exactement ce qui est tapé (attention aux claviers AZERTY "
                                      "pour les chiffres, qui nécessitent la touche Maj).")
        else:
            expected = auth.get_effective_user_password(cfg)
            if pwd == expected:
                self.app.role = "user"
                self.app.show_main()
            else:
                messagebox.showerror("Connexion refusée",
                                      "Mot de passe utilisateur incorrect ou expiré "
                                      "(il change chaque mois).")


# ==========================================================================
# ÉCRAN PRINCIPAL
# ==========================================================================

class MainScreen(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self.pack(fill="both", expand=True)

        top = ttk.Frame(self)
        top.pack(fill="x")
        role_label = "Administrateur" if app.role == "admin" else "Utilisateur"
        ttk.Label(top, text=f"{APP_TITLE}  —  connecté en tant que {role_label}",
                  font=("Segoe UI", 11, "bold")).pack(side="left", padx=10, pady=8)
        ttk.Button(top, text="Se déconnecter", command=app.show_login).pack(side="right", padx=10, pady=8)

        banner = tk.Label(self, text=PAID_SOFTWARE_NOTICE, fg="white", bg="#b8860b",
                           font=("Segoe UI", 9, "bold"), pady=4)
        banner.pack(fill="x")

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.employees_tab = EmployeesTab(notebook, app)
        notebook.add(self.employees_tab, text="Saisie des employés")

        self.payroll_tab = PayrollTab(notebook, app, self.employees_tab)
        notebook.add(self.payroll_tab, text="Bulletins / État de paie")

        self.accounting_tab = AccountingTab(notebook, app, self.payroll_tab)
        notebook.add(self.accounting_tab, text="Écritures comptables")

        self.simulator_tab = SimulatorTab(notebook, app, self.employees_tab)
        notebook.add(self.simulator_tab, text="Simulateur de bulletin")

        if app.role == "admin":
            self.params_tab = ParamsTab(notebook, app)
            notebook.add(self.params_tab, text="Paramètres de paie")

            self.security_tab = SecurityTab(notebook, app)
            notebook.add(self.security_tab, text="Sécurité / Mots de passe")


# ==========================================================================
# ONGLET SAISIE DES EMPLOYÉS
# ==========================================================================

COLUMNS = [
    ("numero", "N°", 40),
    ("nom_prenoms", "Nom & Prénoms", 160),
    ("periode_aff", "Période de paie", 110),
    ("classification", "Classif.", 70),
    ("salaire_base", "Sal. Base", 85),
    ("prime_anciennete", "Prime anc.", 80),
    ("heures_sup", "Heures Sup", 80),
    ("sursalaire", "Sursalaire", 80),
    ("gratification", "Gratif.", 80),
    ("indemnite_caisse", "Indem. Caisse", 90),
    ("indemnite_logement", "Indem. Log.", 90),
    ("indemnite_fonction", "Indem. Fct.", 90),
    ("indemnite_transport", "Indem. Trprt", 90),
    ("personnes_a_charge", "Pers. charge", 85),
    ("retenue_pret", "Retenue prêt", 90),
    ("date_saisie", "Date de saisie", 100),
]

MONEY_COLUMNS = {
    "salaire_base", "prime_anciennete", "heures_sup", "sursalaire", "gratification",
    "indemnite_caisse", "indemnite_logement", "indemnite_fonction", "indemnite_transport",
    "retenue_pret",
}


class EmployeesTab(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app

        # --- Barre d'import en masse (Excel/CSV), utile quand il y a beaucoup
        # d'employés à saisir : on remplit un fichier plutôt que le formulaire.
        toolbar = ttk.Frame(self)
        toolbar.pack(side="top", fill="x", padx=6, pady=(6, 0))
        ttk.Label(toolbar, text="Saisie volumineuse :", font=("Segoe UI", 9, "bold")).pack(side="left")
        ttk.Button(toolbar, text="Importer depuis Excel/CSV",
                   command=self.import_from_file).pack(side="left", padx=(8, 4))
        ttk.Button(toolbar, text="Télécharger le modèle Excel",
                   command=self.download_template).pack(side="left", padx=4)

        # IMPORTANT : on réserve d'abord la place du panneau de droite (largeur
        # fixe) AVANT de placer le tableau (qui a beaucoup de colonnes et
        # utilise fill="both", expand=True). Si on faisait l'inverse, le
        # tableau engloutirait toute la largeur de la fenêtre et le panneau
        # de saisie serait poussé hors champ (invisible), même s'il existe
        # bien dans le code.
        right = ttk.Frame(self, width=340)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)  # garde la largeur réservée même si le contenu est plus petit

        left = ttk.Frame(self)
        left.pack(side="left", fill="both", expand=True, padx=(0, 6))

        cols = [c[0] for c in COLUMNS]
        self.tree = ttk.Treeview(left, columns=cols, show="headings", height=20)
        for key, label, width in COLUMNS:
            self.tree.heading(key, text=label)
            self.tree.column(key, width=width, anchor="center")
        self.tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        left.grid_rowconfigure(0, weight=1)
        left.grid_columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        ttk.Label(right, text="Fiche employé", font=("Segoe UI", 11, "bold")).pack(pady=(4, 10))

        self.form_vars = {}
        form = ttk.Frame(right)
        form.pack(fill="x", padx=8)

        fields = [
            ("nom_prenoms", "Nom & Prénoms", "text"),
            ("periode", "Période de paie (MM/AAAA)", "period"),
            ("classification", "Classification", "combo"),
            ("salaire_base", "Salaire de base", "num"),
            ("prime_anciennete", "Prime d'ancienneté", "num"),
            ("heures_sup", "Heures supplémentaires", "num"),
            ("sursalaire", "Sursalaire", "num"),
            ("gratification", "Gratification", "num"),
            ("indemnite_caisse", "Indemnité Caisse", "num"),
            ("indemnite_logement", "Indemnité Logement", "num"),
            ("indemnite_fonction", "Indemnité Fonction", "num"),
            ("indemnite_transport", "Indemnité Transport", "num"),
            ("personnes_a_charge", "Personnes à charge", "num"),
            ("retenue_pret", "Retenue prêt/avance", "num"),
        ]
        for i, (key, label, kind) in enumerate(fields):
            ttk.Label(form, text=label).grid(row=i, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            if kind == "combo":
                w = ttk.Combobox(form, textvariable=var, values=["CADRE", "AUTRE"],
                                  state="readonly", width=18)
                var.set("AUTRE")
            elif kind == "period":
                w = ttk.Entry(form, textvariable=var, width=20)
                today = datetime.date.today()
                var.set(f"{today.month:02d}/{today.year:04d}")
            else:
                w = ttk.Entry(form, textvariable=var, width=20)
                if kind == "num":
                    var.set("0")
            w.grid(row=i, column=1, pady=2, sticky="w")
            self.form_vars[key] = var

        btns = ttk.Frame(right)
        btns.pack(pady=14)
        ttk.Button(btns, text="Ajouter", command=self.add_employee).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Mettre à jour", command=self.update_employee).grid(row=0, column=1, padx=4)
        ttk.Button(btns, text="Supprimer", command=self.delete_employee).grid(row=0, column=2, padx=4)
        ttk.Button(right, text="Vider le formulaire", command=self.clear_form).pack()

        self.selected_numero = None
        self.selected_date_saisie = None
        self.refresh_tree()

    # ------------------------------------------------------------------
    def get_employees(self):
        return self.app.config_data["employees"]

    def refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for emp in self.get_employees():
            values = []
            for key, _, _ in COLUMNS:
                if key == "periode_aff":
                    values.append(format_period(emp.get("periode", "")))
                elif key in MONEY_COLUMNS:
                    values.append(fmt_amount(emp.get(key, "")))
                else:
                    values.append(emp.get(key, ""))
            self.tree.insert("", "end", iid=str(emp["numero"]), values=values)

    def _read_form(self):
        v = self.form_vars
        try:
            emp = Employee(
                numero=self.selected_numero or self.app.config_data["next_numero"],
                nom_prenoms=v["nom_prenoms"].get().strip(),
                classification=v["classification"].get() or "AUTRE",
                periode=normalize_period(v["periode"].get()),
                salaire_base=float(v["salaire_base"].get() or 0),
                prime_anciennete=float(v["prime_anciennete"].get() or 0),
                heures_sup=float(v["heures_sup"].get() or 0),
                sursalaire=float(v["sursalaire"].get() or 0),
                gratification=float(v["gratification"].get() or 0),
                indemnite_caisse=float(v["indemnite_caisse"].get() or 0),
                indemnite_logement=float(v["indemnite_logement"].get() or 0),
                indemnite_fonction=float(v["indemnite_fonction"].get() or 0),
                indemnite_transport=float(v["indemnite_transport"].get() or 0),
                personnes_a_charge=int(float(v["personnes_a_charge"].get() or 0)),
                retenue_pret=float(v["retenue_pret"].get() or 0),
                date_saisie=self.selected_date_saisie or datetime.date.today().isoformat(),
            )
        except ValueError:
            messagebox.showerror("Erreur de saisie", "Merci de vérifier les valeurs numériques saisies.")
            return None
        if not emp.nom_prenoms:
            messagebox.showerror("Erreur de saisie", "Le nom de l'employé est obligatoire.")
            return None
        return emp

    def _save_or_report(self):
        """Enregistre les données locales ; affiche une erreur claire en cas
        d'échec (ex: permissions) au lieu de laisser l'action passer inaperçue."""
        try:
            storage.save(self.app.config_data)
            return True
        except Exception as exc:
            messagebox.showerror(
                "Impossible d'enregistrer",
                "L'enregistrement local a échoué. Vérifiez que l'application "
                "peut écrire dans votre dossier utilisateur (lancez-la en tant "
                "qu'administrateur, ou vérifiez qu'un antivirus/Windows Defender "
                "ne la bloque pas).\n\n"
                f"Détail technique : {exc}")
            return False

    def add_employee(self):
        emp = self._read_form()
        if emp is None:
            return
        emp.numero = self.app.config_data["next_numero"]
        self.app.config_data["employees"].append(emp.to_dict())
        self.app.config_data["next_numero"] += 1
        if not self._save_or_report():
            # on annule l'ajout en mémoire si l'enregistrement a échoué,
            # pour ne pas désynchroniser l'affichage et le fichier de données
            self.app.config_data["employees"].pop()
            self.app.config_data["next_numero"] -= 1
            return
        self.refresh_tree()
        self.clear_form()
        messagebox.showinfo("Ajouté", f"Employé « {emp.nom_prenoms} » ajouté avec succès.")

    def update_employee(self):
        if self.selected_numero is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé dans la liste.")
            return
        emp = self._read_form()
        if emp is None:
            return
        employees = self.app.config_data["employees"]
        for i, e in enumerate(employees):
            if e["numero"] == self.selected_numero:
                employees[i] = emp.to_dict()
                break
        if not self._save_or_report():
            return
        self.refresh_tree()

    def delete_employee(self):
        if self.selected_numero is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé dans la liste.")
            return
        if not messagebox.askyesno("Confirmer", "Supprimer cet employé ?"):
            return
        employees = self.app.config_data["employees"]
        self.app.config_data["employees"] = [e for e in employees if e["numero"] != self.selected_numero]
        if not self._save_or_report():
            return
        self.refresh_tree()
        self.clear_form()

    def clear_form(self):
        self.selected_numero = None
        self.selected_date_saisie = None
        for key, var in self.form_vars.items():
            if key == "classification":
                var.set("AUTRE")
            elif key == "periode":
                today = datetime.date.today()
                var.set(f"{today.month:02d}/{today.year:04d}")
            elif key == "nom_prenoms":
                var.set("")
            else:
                var.set("0")
        self.tree.selection_remove(self.tree.selection())

    def on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        numero = int(sel[0])
        self.selected_numero = numero
        emp = next((e for e in self.get_employees() if e["numero"] == numero), None)
        if not emp:
            return
        self.selected_date_saisie = emp.get("date_saisie", "")
        for key, var in self.form_vars.items():
            if key == "periode":
                var.set(self._periode_to_input(emp.get("periode", "")))
            elif key in MONEY_COLUMNS or key == "personnes_a_charge":
                # champ modifiable : pas de séparateur de milliers (garde une
                # valeur ré-éditable/parsable), juste sans ".0" superflu
                val = emp.get(key, 0)
                try:
                    f = float(val)
                    var.set(str(int(f)) if f == int(f) else str(f))
                except (TypeError, ValueError):
                    var.set(str(val))
            else:
                var.set(str(emp.get(key, "")))

    @staticmethod
    def _periode_to_input(period_key):
        """Convertit 'AAAA-MM' (stockage) vers 'MM/AAAA' (saisie)."""
        try:
            year, month = period_key.split("-")
            return f"{int(month):02d}/{year}"
        except Exception:
            today = datetime.date.today()
            return f"{today.month:02d}/{today.year:04d}"

    # ------------------------------------------------------------------
    # IMPORT EN MASSE (Excel / CSV)
    # ------------------------------------------------------------------

    # En-têtes reconnus dans le fichier importé -> champ interne de l'employé.
    # Plusieurs variantes acceptées pour plus de souplesse (accents, casse
    # ignorés à la comparaison).
    IMPORT_HEADER_MAP = {
        "nom & prenoms": "nom_prenoms", "nom et prenoms": "nom_prenoms",
        "nom & prénoms": "nom_prenoms", "nom prenoms": "nom_prenoms", "nom": "nom_prenoms",
        "periode de paie": "periode", "période de paie": "periode", "periode": "periode", "mois": "periode",
        "classification": "classification", "classif": "classification", "classif.": "classification",
        "salaire de base": "salaire_base", "sal de base": "salaire_base", "sal. base": "salaire_base",
        "prime d'anciennete": "prime_anciennete", "prime anciennete": "prime_anciennete",
        "prime anc.": "prime_anciennete", "prim anc": "prime_anciennete",
        "heures supplementaires": "heures_sup", "heures sup": "heures_sup", "heure sup": "heures_sup",
        "sursalaire": "sursalaire",
        "gratification": "gratification", "gratif.": "gratification", "gratif": "gratification",
        "indemnite caisse": "indemnite_caisse", "indem. caisse": "indemnite_caisse", "caisse": "indemnite_caisse",
        "indemnite logement": "indemnite_logement", "indem. log.": "indemnite_logement", "log": "indemnite_logement",
        "indemnite fonction": "indemnite_fonction", "indem. fct.": "indemnite_fonction", "fct": "indemnite_fonction",
        "indemnite transport": "indemnite_transport", "indem. trprt": "indemnite_transport", "trprt": "indemnite_transport",
        "personnes a charge": "personnes_a_charge", "pers. charge": "personnes_a_charge", "charges": "personnes_a_charge",
        "retenue pret/avance": "retenue_pret", "retenue pret": "retenue_pret", "retenue prêt": "retenue_pret",
    }

    @staticmethod
    def _normalize_header(text):
        import unicodedata
        text = str(text or "").strip().lower()
        text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
        return text

    def _parse_rows(self, path):
        """Lit un fichier .xlsx ou .csv et retourne une liste de dicts
        {champ_interne: valeur_brute}, à partir de la ligne d'en-tête."""
        ext = path.lower().rsplit(".", 1)[-1]
        header_map = {self._normalize_header(k): v for k, v in self.IMPORT_HEADER_MAP.items()}

        if ext in ("xlsx", "xlsm"):
            import openpyxl
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
            except Exception as exc:
                raise ValueError(
                    "Impossible d'ouvrir ce fichier Excel.\n"
                    "Vérifiez qu'il est bien au format .xlsx (Excel 2007 ou plus récent — "
                    "pas l'ancien .xls), et qu'il n'est pas ouvert dans Excel en ce moment.\n\n"
                    f"Détail : {exc}")

            # On cherche la feuille qui contient les bons en-têtes, plutôt que
            # de se fier uniquement à la feuille "active" (qui peut être une
            # autre feuille selon la dernière feuille consultée dans Excel).
            best_rows, best_score = None, -1
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header = rows[0]
                score = sum(1 for h in header if self._normalize_header(h) in header_map)
                if score > best_score:
                    best_rows, best_score = rows, score
            all_rows = best_rows or []
        elif ext == "xls":
            raise ValueError(
                "Ce fichier est au format Excel 97-2003 (.xls), pas pris en charge.\n"
                "Ouvrez-le dans Excel puis faites « Fichier > Enregistrer sous » "
                "et choisissez le type « Classeur Excel (*.xlsx) », puis réimportez ce nouveau fichier.")
        elif ext == "csv":
            import csv
            # Les CSV exportés par Excel en français sont souvent encodés en
            # Windows-1252 (cp1252) ou Latin-1, pas en UTF-8 : on essaie
            # plusieurs encodages avant d'abandonner.
            raw_bytes = open(path, "rb").read()
            text = None
            for encoding in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    text = raw_bytes.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                raise ValueError("Impossible de lire l'encodage de ce fichier CSV.")

            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,")
            except csv.Error:
                dialect = csv.excel
                dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.reader(text.splitlines(), dialect)
            all_rows = [tuple(row) for row in reader]
        else:
            raise ValueError("Format non pris en charge. Utilisez un fichier .xlsx (Excel) ou .csv.")

        if not all_rows:
            return []

        header_row = all_rows[0]
        field_by_col = {}
        for idx, h in enumerate(header_row):
            norm = self._normalize_header(h)
            if norm in header_map:
                field_by_col[idx] = header_map[norm]

        if "nom_prenoms" not in field_by_col.values():
            raise ValueError("Colonne obligatoire manquante : « Nom & Prénoms ».\n"
                              "Utilisez le bouton « Télécharger le modèle Excel » pour avoir "
                              "les bons en-têtes.")

        records = []
        for row in all_rows[1:]:
            if row is None or all(c in (None, "") for c in row):
                continue
            rec = {}
            for idx, field in field_by_col.items():
                if idx < len(row):
                    rec[field] = row[idx]
            records.append(rec)
        return records

    def import_from_file(self):
        path = filedialog.askopenfilename(
            title="Importer des employés",
            filetypes=[("Excel / CSV", "*.xlsx *.xlsm *.csv"), ("Tous les fichiers", "*.*")],
        )
        if not path:
            return

        try:
            records = self._parse_rows(path)
        except Exception as exc:
            messagebox.showerror("Import impossible", str(exc))
            return

        if not records:
            messagebox.showinfo("Import", "Aucune ligne exploitable trouvée dans ce fichier.")
            return

        def to_float(v):
            if v is None or v == "":
                return 0.0
            if isinstance(v, str):
                v = v.strip()
                # tolère les espaces (séparateur de milliers), le symbole
                # FCFA/CFA, et la virgule décimale française
                v = (v.replace("\xa0", "").replace(" ", "")
                      .replace("FCFA", "").replace("CFA", "").replace("F", "")
                      .replace(",", "."))
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        imported, skipped = 0, 0
        for rec in records:
            nom = str(rec.get("nom_prenoms", "")).strip()
            if not nom:
                skipped += 1
                continue
            classification = str(rec.get("classification", "AUTRE") or "AUTRE").strip().upper()
            if classification not in ("CADRE", "AUTRE"):
                classification = "AUTRE"
            periode = normalize_period(rec.get("periode", ""))
            emp = Employee(
                numero=self.app.config_data["next_numero"],
                nom_prenoms=nom,
                classification=classification,
                periode=periode,
                salaire_base=to_float(rec.get("salaire_base")),
                prime_anciennete=to_float(rec.get("prime_anciennete")),
                heures_sup=to_float(rec.get("heures_sup")),
                sursalaire=to_float(rec.get("sursalaire")),
                gratification=to_float(rec.get("gratification")),
                indemnite_caisse=to_float(rec.get("indemnite_caisse")),
                indemnite_logement=to_float(rec.get("indemnite_logement")),
                indemnite_fonction=to_float(rec.get("indemnite_fonction")),
                indemnite_transport=to_float(rec.get("indemnite_transport")),
                personnes_a_charge=int(to_float(rec.get("personnes_a_charge"))),
                retenue_pret=to_float(rec.get("retenue_pret")),
                date_saisie=datetime.date.today().isoformat(),
            )
            self.app.config_data["employees"].append(emp.to_dict())
            self.app.config_data["next_numero"] += 1
            imported += 1

        storage.save(self.app.config_data)
        self.refresh_tree()
        msg = f"{imported} employé(s) importé(s)."
        if skipped:
            msg += f"\n{skipped} ligne(s) ignorée(s) (nom manquant)."
        messagebox.showinfo("Import terminé", msg)

    def download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Module manquant",
                                  "Le module 'openpyxl' n'est pas installé.\n"
                                  "Installez-le avec : pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_import_employes.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employés"
        headers = ["Nom & Prénoms", "Période de paie (MM/AAAA)", "Classification", "Salaire de base",
                   "Prime d'ancienneté", "Heures supplémentaires", "Sursalaire", "Gratification",
                   "Indemnité Caisse", "Indemnité Logement", "Indemnité Fonction", "Indemnité Transport",
                   "Personnes à charge", "Retenue prêt/avance"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        today = datetime.date.today()
        ws.append(["KABORE Awa", f"{today.month:02d}/{today.year:04d}", "AUTRE",
                   150000, 5000, 0, 0, 0, 10000, 30000, 15000, 20000, 2, 0])
        for i, col in enumerate(ws.columns, start=1):
            length = max((len(str(c.value)) for c in col if c.value is not None), default=12)
            ws.column_dimensions[get_column_letter(i)].width = max(14, length + 2)
        wb.save(path)
        messagebox.showinfo(
            "Modèle créé",
            f"Modèle enregistré :\n{path}\n\n"
            "Remplissez une ligne par employé (la classification doit être "
            "CADRE ou AUTRE), puis utilisez « Importer depuis Excel/CSV ».")


# ==========================================================================
# ONGLET BULLETINS / ÉTAT DE PAIE
# ==========================================================================

class PayrollTab(ttk.Frame):
    def __init__(self, parent, app: App, employees_tab: EmployeesTab):
        super().__init__(parent)
        self.app = app
        self.employees_tab = employees_tab

        top = ttk.Frame(self)
        top.pack(fill="x", pady=6, padx=6)

        today = datetime.date.today()
        ttk.Label(top, text="Période de paie :").pack(side="left")
        self.mois_var = tk.StringVar(value=MOIS_FR[today.month - 1])
        ttk.Combobox(top, textvariable=self.mois_var, values=MOIS_FR, state="readonly",
                     width=12).pack(side="left", padx=6)
        self.annee_var = tk.StringVar(value=str(today.year))
        ttk.Entry(top, textvariable=self.annee_var, width=6).pack(side="left")

        self.all_periods_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text="Toutes périodes confondues",
                         variable=self.all_periods_var).pack(side="left", padx=(10, 0))

        ttk.Button(top, text="Calculer la paie", command=self.calculate).pack(side="left", padx=16)
        ttk.Button(top, text="Exporter vers Excel", command=self.export_excel).pack(side="left", padx=(0, 6))
        ttk.Button(top, text="Bulletin PDF (sélection)", command=self.export_selected_payslip).pack(side="left", padx=(0, 6))
        ttk.Button(top, text="Tous les bulletins (PDF)", command=self.export_all_payslips).pack(side="left")

        result_cols = ["numero", "nom_prenoms", "remuneration_totale", "cnss_salariale",
                        "salaire_brut", "base_imposable", "iuts_net", "salaire_net",
                        "retenue_obligatoire", "retenue_pret", "net_percu", "cout_total_employeur"]
        headers = ["N°", "Nom & Prénoms", "Rém. Totale", "CNSS", "Sal. Brut",
                   "Base Imposable", "IUTS", "Salaire Net", "Ret. Oblig. 1%",
                   "Ret. Prêt", "Net Perçu", "Coût Employeur"]

        self.result_cols = result_cols
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=6)
        self.tree = ttk.Treeview(tree_frame, columns=result_cols, show="headings", height=20)
        for key, label in zip(result_cols, headers):
            self.tree.heading(key, text=label)
            self.tree.column(key, width=105, anchor="center")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        totals = ttk.Frame(self)
        totals.pack(fill="x", padx=6, pady=(0, 6))
        self.totals_label = ttk.Label(totals, text="", font=("Segoe UI", 10, "bold"))
        self.totals_label.pack(side="left")

        self.last_results = []

    def selected_period_key(self):
        month = MOIS_FR.index(self.mois_var.get()) + 1
        try:
            year = int(self.annee_var.get())
        except ValueError:
            year = datetime.date.today().year
        return f"{year:04d}-{month:02d}"

    def calculate(self):
        params = self.app.config_data["params"]
        employees = self.app.config_data["employees"]
        if not self.all_periods_var.get():
            period_key = self.selected_period_key()
            employees = [e for e in employees if e.get("periode") == period_key]
        self.tree.delete(*self.tree.get_children())
        results = []
        total_net = total_cnss = total_iuts = total_cout = total_ro = 0.0
        for e in employees:
            emp = Employee(**e)
            r = compute_payslip(emp, params)
            results.append(r)
            values = [r["numero"] if k == "numero" else
                      r["nom_prenoms"] if k == "nom_prenoms" else
                      fmt_amount(r[k]) for k in self.result_cols]
            self.tree.insert("", "end", values=values)
            total_net += r["net_percu"]
            total_cnss += r["cnss_total"]
            total_iuts += r["iuts_net"]
            total_cout += r["cout_total_employeur"]
            total_ro += r["retenue_obligatoire"]
        self.last_results = results
        self.totals_label.config(
            text=(f"Total Net Perçu : {total_net:,.0f}  |  Total CNSS : {total_cnss:,.0f}  |  "
                  f"Total IUTS : {total_iuts:,.0f}  |  Total Ret. Oblig. : {total_ro:,.0f}  |  "
                  f"Coût total employeur : {total_cout:,.0f}  FCFA")
            .replace(",", " ")
        )
        if not employees:
            messagebox.showinfo("Info", "Aucun employé saisi pour le moment.")

    def export_excel(self):
        if not self.last_results:
            self.calculate()
        if not self.last_results:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Module manquant",
                                  "Le module 'openpyxl' n'est pas installé.\n"
                                  "Installez-le avec : pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"Etat_de_paie_{self.mois_var.get()}_{self.annee_var.get()}.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "État de paie"

        title = f"ÉTAT DE PAIE — {self.mois_var.get().upper()} {self.annee_var.get()}"
        ws.merge_cells("A1:K1")
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)

        headers = ["N°", "Nom & Prénoms", "Classification", "Rém. Totale", "CNSS Salariale",
                   "Salaire Brut", "Base Imposable", "IUTS Net", "Salaire Net",
                   "Retenue Obligatoire 1%", "Retenue Prêt", "Net Perçu", "Coût Total Employeur"]
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")

        for r in self.last_results:
            ws.append([
                r["numero"], r["nom_prenoms"], r["classification"], r["remuneration_totale"],
                r["cnss_salariale"], r["salaire_brut"], r["base_imposable"], r["iuts_net"],
                r["salaire_net"], r["retenue_obligatoire"], r["retenue_pret"], r["net_percu"],
                r["cout_total_employeur"],
            ])

        for i, col in enumerate(ws.columns, start=1):
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(i)].width = max(12, length + 2)

        notice_row = ws.max_row + 2
        ws.cell(row=notice_row, column=1, value=PAID_SOFTWARE_NOTICE).font = Font(italic=True, color="B8860B")

        wb.save(path)
        messagebox.showinfo("Export réussi", f"Fichier exporté :\n{path}")

    # ------------------------------------------------------------------
    # BULLETINS DE PAIE (PDF), avec en-tête et pied de page paramétrables
    # ------------------------------------------------------------------

    def _period_display(self):
        return f"{self.mois_var.get()} {self.annee_var.get()}"

    def export_selected_payslip(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé dans le tableau ci-dessus "
                                         "(après avoir cliqué sur « Calculer la paie »).")
            return
        if not self.last_results:
            messagebox.showinfo("Info", "Cliquez d'abord sur « Calculer la paie ».")
            return
        values = self.tree.item(sel[0], "values")
        numero = int(values[0])
        result = next((r for r in self.last_results if r["numero"] == numero), None)
        if result is None:
            messagebox.showerror("Erreur", "Employé introuvable dans les résultats calculés.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Bulletin_{result['nom_prenoms'].replace(' ', '_')}_{self.mois_var.get()}_{self.annee_var.get()}.pdf",
        )
        if not path:
            return
        try:
            self._generate_payslips_pdf([result], path)
        except ImportError:
            messagebox.showerror("Module manquant",
                                  "Le module 'reportlab' n'est pas installé.\n"
                                  "Installez-le avec : pip install reportlab")
            return
        messagebox.showinfo("Export réussi", f"Bulletin de paie généré :\n{path}")

    def export_all_payslips(self):
        if not self.last_results:
            self.calculate()
        if not self.last_results:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Bulletins_de_paie_{self.mois_var.get()}_{self.annee_var.get()}.pdf",
        )
        if not path:
            return
        try:
            self._generate_payslips_pdf(self.last_results, path)
        except ImportError:
            messagebox.showerror("Module manquant",
                                  "Le module 'reportlab' n'est pas installé.\n"
                                  "Installez-le avec : pip install reportlab")
            return
        messagebox.showinfo("Export réussi",
                             f"{len(self.last_results)} bulletin(s) de paie généré(s) dans :\n{path}")

    def _generate_payslips_pdf(self, results, path):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as pdf_canvas

        entete = self.app.config_data.get("bulletin_entete", {}) or {}
        pied = self.app.config_data.get("bulletin_pied_de_page", "") or ""
        periode_txt = self._period_display()

        width, height = A4
        c = pdf_canvas.Canvas(path, pagesize=A4)

        for result in results:
            self._draw_payslip_page(c, width, height, mm, entete, pied, periode_txt, result)
            c.showPage()

        c.save()

    def _draw_payslip_page(self, c, width, height, mm, entete, pied, periode_txt, r):
        x_left = 18 * mm
        x_right = width - 18 * mm
        y = height - 18 * mm

        # --- En-tête ---------------------------------------------------
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x_left, y, entete.get("nom_entreprise") or "Mon Entreprise")
        y -= 6 * mm
        c.setFont("Helvetica", 9)
        coords = [v for v in (entete.get("adresse"), entete.get("telephone"), entete.get("email")) if v]
        if coords:
            c.drawString(x_left, y, "  •  ".join(coords))
            y -= 5 * mm
        if entete.get("note_entete"):
            c.setFont("Helvetica-Oblique", 8)
            c.drawString(x_left, y, entete["note_entete"])
            y -= 5 * mm

        y -= 2 * mm
        c.setLineWidth(1)
        c.line(x_left, y, x_right, y)
        y -= 8 * mm

        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(width / 2, y, f"BULLETIN DE PAIE — {periode_txt.upper()}")
        y -= 10 * mm

        # --- Bloc employé -----------------------------------------------
        c.setFont("Helvetica", 10)
        c.drawString(x_left, y, f"N° employé : {r['numero']}")
        c.drawString(width / 2, y, f"Classification : {r['classification']}")
        y -= 6 * mm
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x_left, y, f"{r['nom_prenoms']}")
        y -= 5 * mm
        c.setFont("Helvetica", 10)
        c.drawString(x_left, y, f"Personnes à charge : {r['personnes_a_charge']}")
        y -= 9 * mm

        def money(v):
            return f"{v:,.0f}".replace(",", " ") + " FCFA"

        def row(label, value, y, bold=False, indent=0):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", 10)
            c.drawString(x_left + indent, y, label)
            c.drawRightString(x_right, y, money(value))
            return y - 5.5 * mm

        c.setLineWidth(0.7)
        c.line(x_left, y, x_right, y)
        y -= 6 * mm
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x_left, y, "GAINS")
        y -= 6 * mm

        gain_lines = [
            ("Salaire de base", r["salaire_base"]),
            ("Prime d'ancienneté", r["prime_anciennete"]),
            ("Heures supplémentaires", r["heures_sup"]),
            ("Sursalaire", r["sursalaire"]),
            ("Gratification", r["gratification"]),
            ("Indemnité de caisse", r["indemnite_caisse"]),
            ("Indemnité de logement", r["indemnite_logement"]),
            ("Indemnité de fonction", r["indemnite_fonction"]),
            ("Indemnité de transport", r["indemnite_transport"]),
        ]
        for label, value in gain_lines:
            if value:  # on n'affiche pas les lignes à 0, pour un bulletin plus lisible
                y = row(label, value, y, indent=2 * mm)
        y -= 1 * mm
        c.setLineWidth(0.4)
        c.line(x_left + 2 * mm, y, x_right, y)
        y -= 5.5 * mm
        y = row("Rémunération totale", r["remuneration_totale"], y, bold=True)
        y -= 3 * mm

        c.setFont("Helvetica-Bold", 10)
        c.drawString(x_left, y, "RETENUES")
        y -= 6 * mm
        y = row("CNSS (part salariale)", r["cnss_salariale"], y)
        y = row("IUTS", r["iuts_net"], y)
        y = row("Retenue obligatoire (1%)", r["retenue_obligatoire"], y)
        y = row("Retenue prêt / avance", r["retenue_pret"], y)
        y -= 3 * mm

        c.setLineWidth(1)
        c.line(x_left, y, x_right, y)
        y -= 8 * mm
        c.setFont("Helvetica-Bold", 13)
        c.drawString(x_left, y, "NET À PAYER")
        c.drawRightString(x_right, y, money(r["net_percu"]))
        y -= 8 * mm
        c.setLineWidth(1)
        c.line(x_left, y, x_right, y)
        y -= 10 * mm

        c.setFont("Helvetica", 9)
        c.drawString(x_left, y, f"Coût total employeur (charges patronales incluses) : {money(r['cout_total_employeur'])}")

        # --- Pied de page -------------------------------------------------
        bottom = 30 * mm
        c.setLineWidth(0.5)
        c.line(x_left, bottom + 14 * mm, x_right, bottom + 14 * mm)
        c.setFont("Helvetica", 9)
        c.drawString(x_left, bottom + 8 * mm, "Signature de l'employeur")
        c.drawRightString(x_right, bottom + 8 * mm, "Signature de l'employé")

        if pied:
            c.setFont("Helvetica-Oblique", 7.5)
            text_obj = c.beginText(x_left, bottom)
            text_obj.setLeading(9)
            for line in pied.split("\n"):
                text_obj.textLine(line)
            c.drawText(text_obj)


# Génère l'écriture de paie en partie double (Débit / Crédit), à partir des
# résultats calculés dans l'onglet "Bulletins / État de paie". Comptes basés
# sur le plan comptable SYSCOHADA habituellement utilisé pour la paie.

ACCOUNTING_ACCOUNTS = {
    "salaire_base": ("661100", "Salaires de base"),
    "primes_gratif": ("661200", "Primes d'ancienneté et gratifications"),
    "heures_sursal": ("661800", "Heures supplémentaires et sursalaire"),
    "indemnite_caisse": ("663800", "Indemnité de caisse"),
    "indemnite_logement": ("663100", "Indemnité de logement"),
    "indemnite_fonction": ("663200", "Indemnité de fonction"),
    "indemnite_transport": ("663400", "Indemnité de transport"),
    "salaire_net": ("422000", "Salaires nets à payer"),
    "retenue_obligatoire": ("447220", "Retenue obligatoire (RO) sur salaires"),
    "retenue_pret": ("421000", "Retenues sur avances / prêts au personnel"),
    "cnss_salariale": ("431300", "CNSS — part salariale"),
    "iuts_net": ("447210", "IUTS retenu à la source"),
    "cnss_patronale": ("664100", "Charges sociales — CNSS patronale"),
    "cnss_patronale_credit": ("431300", "CNSS — part patronale"),
    "tpa": ("664200", "Taxe patronale d'apprentissage (TPA)"),
    "tpa_credit": ("447230", "TPA à reverser"),
}


class AccountingTab(ttk.Frame):
    def __init__(self, parent, app: App, payroll_tab: "PayrollTab"):
        super().__init__(parent)
        self.app = app
        self.payroll_tab = payroll_tab

        top = ttk.Frame(self)
        top.pack(fill="x", padx=6, pady=6)
        ttk.Label(top, text="Écritures comptables de la paie :",
                  font=("Segoe UI", 11, "bold")).pack(side="left")

        today = datetime.date.today()
        ttk.Label(top, text="  Période :").pack(side="left")
        self.mois_var = tk.StringVar(value=MOIS_FR[today.month - 1])
        ttk.Combobox(top, textvariable=self.mois_var, values=MOIS_FR, state="readonly",
                     width=12).pack(side="left", padx=4)
        self.annee_var = tk.StringVar(value=str(today.year))
        ttk.Entry(top, textvariable=self.annee_var, width=6).pack(side="left")

        self.all_periods_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text="Toutes périodes confondues",
                         variable=self.all_periods_var).pack(side="left", padx=(10, 0))

        ttk.Button(top, text="Générer", command=self.generate).pack(side="left", padx=16)
        ttk.Button(top, text="Exporter vers Excel", command=self.export_excel).pack(side="left")

        cols = ["compte", "libelle", "debit", "credit"]
        headers = ["N° Compte", "Libellé", "Débit", "Crédit"]
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=6)
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=20)
        widths = [100, 380, 130, 130]
        for key, label, w in zip(cols, headers, widths):
            self.tree.heading(key, text=label)
            self.tree.column(key, width=w, anchor="center" if key != "libelle" else "w")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.totals_label = ttk.Label(self, text="", font=("Segoe UI", 10, "bold"))
        self.totals_label.pack(anchor="w", padx=6, pady=(0, 6))

        self.last_rows = []

    def selected_period_key(self):
        month = MOIS_FR.index(self.mois_var.get()) + 1
        try:
            year = int(self.annee_var.get())
        except ValueError:
            year = datetime.date.today().year
        return f"{year:04d}-{month:02d}"

    def _filtered_employees(self):
        raw = self.app.config_data["employees"]
        if not self.all_periods_var.get():
            period_key = self.selected_period_key()
            raw = [e for e in raw if e.get("periode") == period_key]
        return [Employee(**e) for e in raw]

    def _build_rows(self, employees, results):
        def s(key):
            return sum(r[key] for r in results)

        rows = []
        sum_sal_base = sum(e.salaire_base for e in employees)
        sum_primes = sum(e.prime_anciennete + e.gratification for e in employees)
        sum_hs = sum(e.heures_sup + e.sursalaire for e in employees)
        sum_caisse = sum(e.indemnite_caisse for e in employees)
        sum_log = sum(e.indemnite_logement for e in employees)
        sum_fct = sum(e.indemnite_fonction for e in employees)
        sum_trp = sum(e.indemnite_transport for e in employees)

        rows.append(("661100", "SALAIRES DE BASE", sum_sal_base, 0))
        rows.append(("661200", "PRIMES ANCIENNETÉ ET GRATIFICATIONS", sum_primes, 0))
        rows.append(("661800", "HEURES SUPPLÉMENTAIRES ET SURSALAIRE", sum_hs, 0))
        rows.append(("663800", "INDEMNITÉ DE CAISSE", sum_caisse, 0))
        rows.append(("663100", "INDEMNITÉ DE LOGEMENT", sum_log, 0))
        rows.append(("663200", "INDEMNITÉ DE FONCTION", sum_fct, 0))
        rows.append(("663400", "INDEMNITÉ DE TRANSPORT", sum_trp, 0))

        total_net = s("net_percu")
        total_ro = s("retenue_obligatoire")
        total_pret = s("retenue_pret")
        total_cnss_sal = s("cnss_salariale")
        total_iuts = s("iuts_net")

        rows.append(("422000", "SALAIRES NETS À PAYER", 0, total_net))
        rows.append(("447220", "RETENUE OBLIGATOIRE (RO) SUR SALAIRES", 0, total_ro))
        rows.append(("421000", "RETENUES SUR AVANCES / PRÊTS AU PERSONNEL", 0, total_pret))
        rows.append(("431300", "CNSS — PART SALARIALE", 0, total_cnss_sal))
        rows.append(("447210", "IUTS RETENU À LA SOURCE", 0, total_iuts))

        sous_total_1_debit = sum_sal_base + sum_primes + sum_hs + sum_caisse + sum_log + sum_fct + sum_trp
        sous_total_1_credit = total_net + total_ro + total_pret + total_cnss_sal + total_iuts
        rows.append(("", "SOUS-TOTAL 1 (charges de personnel)", sous_total_1_debit, sous_total_1_credit))

        # --- Charges patronales
        total_cnss_pat = s("cnss_patronale")
        total_tpa = s("tpa_patronale")
        rows.append(("664100", "CHARGES SOCIALES — CNSS PATRONALE", total_cnss_pat, 0))
        rows.append(("431300", "CNSS — PART PATRONALE (à reverser)", 0, total_cnss_pat))
        rows.append(("664200", "TAXE PATRONALE D'APPRENTISSAGE (TPA)", total_tpa, 0))
        rows.append(("447230", "TPA À REVERSER", 0, total_tpa))

        sous_total_2 = total_cnss_pat + total_tpa
        rows.append(("", "SOUS-TOTAL 2 (charges patronales)", sous_total_2, sous_total_2))

        grand_total_debit = sous_total_1_debit + sous_total_2
        grand_total_credit = sous_total_1_credit + sous_total_2
        rows.append(("", "GRAND TOTAL", grand_total_debit, grand_total_credit))

        return rows

    def generate(self):
        params = self.app.config_data["params"]
        employees = self._filtered_employees()
        if not employees:
            period_txt = "toutes périodes" if self.all_periods_var.get() else format_period(self.selected_period_key())
            messagebox.showinfo("Info", f"Aucun employé pour la période sélectionnée ({period_txt}).")
            self.tree.delete(*self.tree.get_children())
            self.last_rows = []
            self.totals_label.config(text="")
            return
        results = [compute_payslip(emp, params) for emp in employees]

        rows = self._build_rows(employees, results)
        self.last_rows = rows
        self.tree.delete(*self.tree.get_children())
        for compte, libelle, debit, credit in rows:
            bold = libelle.startswith(("SOUS-TOTAL", "GRAND TOTAL"))
            values = (compte, libelle, f"{debit:,.0f}".replace(",", " ") if debit else "",
                      f"{credit:,.0f}".replace(",", " ") if credit else "")
            iid = self.tree.insert("", "end", values=values, tags=("total",) if bold else ())
        self.tree.tag_configure("total", font=("Segoe UI", 9, "bold"), background="#eef2f7")

        total_debit = rows[-1][2]
        total_credit = rows[-1][3]
        equilibre = "✓ Écriture équilibrée" if abs(total_debit - total_credit) < 1 else "⚠ ÉCRITURE DÉSÉQUILIBRÉE"
        self.totals_label.config(text=f"{equilibre}  —  Total Débit : {total_debit:,.0f}  |  "
                                       f"Total Crédit : {total_credit:,.0f}  FCFA".replace(",", " "))

    def export_excel(self):
        if not self.last_rows:
            self.generate()
        if not self.last_rows:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Module manquant",
                                  "Le module 'openpyxl' n'est pas installé.\n"
                                  "Installez-le avec : pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Ecritures_comptables_paie.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Écritures comptables"

        ws.merge_cells("A1:D1")
        ws["A1"] = "ÉCRITURE COMPTABLE DE PAIE"
        ws["A1"].font = Font(size=14, bold=True)

        ws.append([])
        ws.append(["N° Compte", "Libellé", "Débit", "Crédit"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")

        for compte, libelle, debit, credit in self.last_rows:
            ws.append([compte, libelle, debit or None, credit or None])

        for i, col in enumerate(ws.columns, start=1):
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(i)].width = max(14, length + 2)

        notice_row = ws.max_row + 2
        ws.cell(row=notice_row, column=1, value=PAID_SOFTWARE_NOTICE).font = Font(italic=True, color="B8860B")

        wb.save(path)
        messagebox.showinfo("Export réussi", f"Fichier exporté :\n{path}")


# ==========================================================================
# ONGLET SIMULATEUR DE BULLETIN (net -> base + indemnités)
# ==========================================================================

class SimulatorTab(ttk.Frame):
    def __init__(self, parent, app: App, employees_tab: "EmployeesTab"):
        super().__init__(parent)
        self.app = app
        self.employees_tab = employees_tab
        self.last_result = None
        self.last_employee_template = None

        left = ttk.Frame(self)
        left.pack(side="left", fill="y", padx=10, pady=10)

        ttk.Label(left, text="Simulateur : Net → Salaire de base",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 4))
        ttk.Label(left, text="Entrez le net souhaité, le logiciel retrouve\n"
                              "automatiquement le salaire de base à appliquer.",
                  foreground="#555", justify="left").pack(anchor="w", pady=(0, 12))

        form = ttk.Frame(left)
        form.pack(anchor="w")
        self.vars = {}

        def field(label, key, default, row, kind="num"):
            ttk.Label(form, text=label).grid(row=row, column=0, sticky="w", pady=3)
            var = tk.StringVar(value=str(default))
            if kind == "combo":
                w = ttk.Combobox(form, textvariable=var, values=["CADRE", "AUTRE"],
                                  state="readonly", width=18)
            else:
                w = ttk.Entry(form, textvariable=var, width=20)
            w.grid(row=row, column=1, pady=3, sticky="w")
            self.vars[key] = var
            return w

        field("Nom & Prénoms (optionnel)", "nom_prenoms", "", 0, kind="text")
        field("Classification", "classification", "AUTRE", 1, kind="combo")
        field("Personnes à charge", "personnes_a_charge", "0", 2)
        field("Prime d'ancienneté", "prime_anciennete", "0", 3)
        field("Heures supplémentaires", "heures_sup", "0", 4)
        field("Sursalaire", "sursalaire", "0", 5)
        field("Gratification", "gratification", "0", 6)
        field("Indemnité Caisse", "indemnite_caisse", "0", 7)
        field("Retenue prêt/avance", "retenue_pret", "0", 8)

        self.auto_indem_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            left, text="Optimiser automatiquement les indemnités Logement/\n"
                       "Fonction/Transport (aux plafonds fiscaux exonérés)",
            variable=self.auto_indem_var, command=self._toggle_indem_fields
        ).pack(anchor="w", pady=(10, 4))

        indem_form = ttk.Frame(left)
        indem_form.pack(anchor="w")
        params = self.app.config_data["params"]
        self.indem_entries = {}
        for i, (label, key, param_key) in enumerate([
            ("Indemnité Logement", "indemnite_logement", "exo_logement"),
            ("Indemnité Fonction", "indemnite_fonction", "exo_fonction"),
            ("Indemnité Transport", "indemnite_transport", "exo_transport"),
        ]):
            plafond = params[param_key][1]
            ttk.Label(indem_form, text=label).grid(row=i, column=0, sticky="w", pady=3)
            var = tk.StringVar(value=str(plafond))
            entry = ttk.Entry(indem_form, textvariable=var, width=20, state="readonly")
            entry.grid(row=i, column=1, pady=3, sticky="w")
            self.vars[key] = var
            self.indem_entries[key] = entry

        ttk.Separator(left).pack(fill="x", pady=12)
        ttk.Label(left, text="Net à payer souhaité (FCFA)", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.target_var = tk.StringVar(value="150000")
        ttk.Entry(left, textvariable=self.target_var, width=22, font=("Segoe UI", 11)).pack(anchor="w", pady=(2, 10))

        ttk.Button(left, text="Simuler", command=self.simulate).pack(anchor="w", pady=(0, 6))
        self.create_btn = ttk.Button(left, text="Créer l'employé à partir de cette simulation",
                                      command=self.create_employee, state="disabled")
        self.create_btn.pack(anchor="w")

        # --- Zone de résultat --------------------------------------------
        right = ttk.Frame(self)
        right.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ttk.Label(right, text="Résultat de la simulation", font=("Segoe UI", 12, "bold")).pack(anchor="w")

        self.result_text = tk.Text(right, width=52, height=26, font=("Consolas", 10),
                                    state="disabled", relief="solid", borderwidth=1)
        self.result_text.pack(anchor="w", pady=(8, 0), fill="y")

    def _toggle_indem_fields(self):
        auto = self.auto_indem_var.get()
        params = self.app.config_data["params"]
        mapping = {"indemnite_logement": "exo_logement", "indemnite_fonction": "exo_fonction",
                   "indemnite_transport": "exo_transport"}
        for key, entry in self.indem_entries.items():
            if auto:
                self.vars[key].set(str(params[mapping[key]][1]))
                entry.config(state="readonly")
            else:
                entry.config(state="normal")

    def _read_template(self):
        v = self.vars
        try:
            emp = Employee(
                numero=0,
                nom_prenoms=v["nom_prenoms"].get().strip() or "Simulation",
                classification=v["classification"].get() or "AUTRE",
                periode="",
                salaire_base=0.0,
                prime_anciennete=float(v["prime_anciennete"].get() or 0),
                heures_sup=float(v["heures_sup"].get() or 0),
                sursalaire=float(v["sursalaire"].get() or 0),
                gratification=float(v["gratification"].get() or 0),
                indemnite_caisse=float(v["indemnite_caisse"].get() or 0),
                indemnite_logement=float(v["indemnite_logement"].get() or 0),
                indemnite_fonction=float(v["indemnite_fonction"].get() or 0),
                indemnite_transport=float(v["indemnite_transport"].get() or 0),
                personnes_a_charge=int(float(v["personnes_a_charge"].get() or 0)),
                retenue_pret=float(v["retenue_pret"].get() or 0),
            )
        except ValueError:
            messagebox.showerror("Erreur de saisie", "Merci de vérifier les valeurs numériques saisies.")
            return None
        try:
            target = float(self.target_var.get().replace(" ", "").replace(",", "."))
        except ValueError:
            messagebox.showerror("Erreur de saisie", "Le « Net à payer souhaité » doit être un nombre.")
            return None
        if target <= 0:
            messagebox.showerror("Erreur de saisie", "Le « Net à payer souhaité » doit être positif.")
            return None
        return emp, target

    def simulate(self):
        parsed = self._read_template()
        if parsed is None:
            return
        emp_template, target = parsed
        params = self.app.config_data["params"]

        base, r = find_base_for_target_net(emp_template, params, target_net=target)
        self.last_result = r
        self.last_employee_template = replace(emp_template, salaire_base=base)
        self.create_btn.config(state="normal")

        def money(v):
            return f"{v:,.0f}".replace(",", " ") + " FCFA"

        ecart = r["net_percu"] - target
        lines = [
            f"Net à payer souhaité      : {money(target)}",
            f"Net à payer obtenu        : {money(r['net_percu'])}  (écart : {ecart:+.0f})",
            "",
            f"→ Salaire de base à payer : {money(base)}",
            "",
            "── Détail du bulletin obtenu ──────────────────",
            f"Indemnité Logement         {money(r['indemnite_logement'])}",
            f"Indemnité Fonction         {money(r['indemnite_fonction'])}",
            f"Indemnité Transport        {money(r['indemnite_transport'])}",
            f"Indemnité Caisse           {money(r['indemnite_caisse'])}",
            f"Prime d'ancienneté         {money(r['prime_anciennete'])}",
            f"Heures sup. + Sursalaire   {money(r['heures_sup'] + r['sursalaire'])}",
            f"Gratification              {money(r['gratification'])}",
            "─────────────────────────────────────────────",
            f"Rémunération totale        {money(r['remuneration_totale'])}",
            "",
            f"CNSS (salariale)           {money(r['cnss_salariale'])}",
            f"IUTS                       {money(r['iuts_net'])}",
            f"Retenue obligatoire (1%)   {money(r['retenue_obligatoire'])}",
            f"Retenue prêt/avance        {money(r['retenue_pret'])}",
            "─────────────────────────────────────────────",
            f"NET À PAYER                {money(r['net_percu'])}",
            "",
            f"Coût total employeur       {money(r['cout_total_employeur'])}",
        ]

        self.result_text.config(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", "\n".join(lines))
        self.result_text.config(state="disabled")

    def create_employee(self):
        if self.last_employee_template is None:
            return
        emp = self.last_employee_template
        if not emp.nom_prenoms or emp.nom_prenoms == "Simulation":
            messagebox.showinfo("Nom requis",
                                 "Renseignez le champ « Nom & Prénoms » avant de créer l'employé.")
            return
        emp = replace(emp, numero=self.app.config_data["next_numero"],
                      periode=current_period_key(),
                      date_saisie=datetime.date.today().isoformat())
        self.app.config_data["employees"].append(emp.to_dict())
        self.app.config_data["next_numero"] += 1
        try:
            storage.save(self.app.config_data)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible d'enregistrer : {exc}")
            return
        self.employees_tab.refresh_tree()
        messagebox.showinfo("Employé créé",
                             f"« {emp.nom_prenoms} » a été ajouté à la liste des employés\n"
                             f"(période : {format_period(emp.periode)}).")


# ==========================================================================
# ONGLET PARAMÈTRES (Administrateur uniquement)
# ==========================================================================

class ParamsTab(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        params = app.config_data["params"]

        canvas = tk.Canvas(self, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.vars = {}
        row = 0

        def section(title):
            nonlocal row
            ttk.Label(inner, text=title, font=("Segoe UI", 11, "bold")).grid(
                row=row, column=0, columnspan=2, sticky="w", pady=(14, 4), padx=8)
            row += 1

        def field(label, key, value):
            nonlocal row
            ttk.Label(inner, text=label).grid(row=row, column=0, sticky="w", padx=8, pady=2)
            var = tk.StringVar(value=str(value))
            ttk.Entry(inner, textvariable=var, width=16).grid(row=row, column=1, sticky="w", padx=8, pady=2)
            self.vars[key] = var
            row += 1

        section("1. Cotisations CNSS")
        field("Taux CNSS salarié", "taux_cnss_salarie", params["taux_cnss_salarie"])
        field("Plafond mensuel soumis à CNSS", "plafond_cnss", params["plafond_cnss"])
        field("Cotisation CNSS salariale plafonnée", "cnss_salariale_plafonnee", params["cnss_salariale_plafonnee"])
        field("Taux CNSS patronale", "taux_cnss_patronale", params["taux_cnss_patronale"])
        field("Taux TPA", "taux_tpa", params["taux_tpa"])
        field("Retenue obligatoire sur salaire net", "taux_retenue_obligatoire", params["taux_retenue_obligatoire"])

        section("2. Abattement forfaitaire (IUTS)")
        field("Taux abattement CADRE", "abattement_cadre", params["abattement_cadre"])
        field("Taux abattement AUTRE", "abattement_autre", params["abattement_autre"])

        section("3. Plafond fiscal")
        field("Taux du plafond fiscal", "taux_plafond_fiscal", params["taux_plafond_fiscal"])

        section("4. Exonération des indemnités (taux, plafond)")
        field("Logement — taux exonéré", "exo_logement_taux", params["exo_logement"][0])
        field("Logement — plafond mensuel", "exo_logement_plafond", params["exo_logement"][1])
        field("Fonction — taux exonéré", "exo_fonction_taux", params["exo_fonction"][0])
        field("Fonction — plafond mensuel", "exo_fonction_plafond", params["exo_fonction"][1])
        field("Transport — taux exonéré", "exo_transport_taux", params["exo_transport"][0])
        field("Transport — plafond mensuel", "exo_transport_plafond", params["exo_transport"][1])

        section("Barème IUTS et réduction pour charges de famille")
        ttk.Label(inner, text="(modifiables uniquement dans le fichier de données JSON — "
                               "voir le bouton ci-dessous)", foreground="#666").grid(
            row=row, column=0, columnspan=2, sticky="w", padx=8)
        row += 1

        # --- En-tête / pied de page du bulletin de paie (PDF) ---------------
        section("5. En-tête et pied de page du bulletin de paie (PDF)")
        entete = self.app.config_data.get("bulletin_entete", {})
        self.text_vars = {}

        def text_field(label, key, value, width=40):
            nonlocal row
            ttk.Label(inner, text=label).grid(row=row, column=0, sticky="w", padx=8, pady=2)
            var = tk.StringVar(value=str(value or ""))
            ttk.Entry(inner, textvariable=var, width=width).grid(
                row=row, column=1, sticky="w", padx=8, pady=2)
            self.text_vars[key] = var
            row += 1

        text_field("Nom de l'entreprise (en-tête)", "nom_entreprise", entete.get("nom_entreprise", ""))
        text_field("Adresse", "adresse", entete.get("adresse", ""))
        text_field("Téléphone", "telephone", entete.get("telephone", ""))
        text_field("Email", "email", entete.get("email", ""))
        text_field("Note supplémentaire en en-tête (optionnel)", "note_entete", entete.get("note_entete", ""))

        ttk.Label(inner, text="Texte du pied de page (mentions légales, signature...)").grid(
            row=row, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 2))
        row += 1
        self.footer_text = tk.Text(inner, width=60, height=4, wrap="word")
        self.footer_text.insert("1.0", self.app.config_data.get("bulletin_pied_de_page", ""))
        self.footer_text.grid(row=row, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))
        row += 1

        ttk.Button(inner, text="Ouvrir le dossier des données",
                   command=self.open_data_folder).grid(row=row, column=0, pady=16, padx=8, sticky="w")
        ttk.Button(inner, text="Enregistrer les paramètres",
                   command=self.save_params).grid(row=row, column=1, pady=16, padx=8, sticky="w")

    def save_params(self):
        try:
            p = self.app.config_data["params"]
            v = self.vars
            p["taux_cnss_salarie"] = float(v["taux_cnss_salarie"].get())
            p["plafond_cnss"] = float(v["plafond_cnss"].get())
            p["cnss_salariale_plafonnee"] = float(v["cnss_salariale_plafonnee"].get())
            p["taux_cnss_patronale"] = float(v["taux_cnss_patronale"].get())
            p["taux_tpa"] = float(v["taux_tpa"].get())
            p["taux_retenue_obligatoire"] = float(v["taux_retenue_obligatoire"].get())
            p["abattement_cadre"] = float(v["abattement_cadre"].get())
            p["abattement_autre"] = float(v["abattement_autre"].get())
            p["taux_plafond_fiscal"] = float(v["taux_plafond_fiscal"].get())
            p["exo_logement"] = [float(v["exo_logement_taux"].get()), float(v["exo_logement_plafond"].get())]
            p["exo_fonction"] = [float(v["exo_fonction_taux"].get()), float(v["exo_fonction_plafond"].get())]
            p["exo_transport"] = [float(v["exo_transport_taux"].get()), float(v["exo_transport_plafond"].get())]
        except ValueError:
            messagebox.showerror("Erreur", "Merci de vérifier les valeurs saisies (nombres attendus).")
            return

        self.app.config_data["bulletin_entete"] = {
            "nom_entreprise": self.text_vars["nom_entreprise"].get().strip(),
            "adresse": self.text_vars["adresse"].get().strip(),
            "telephone": self.text_vars["telephone"].get().strip(),
            "email": self.text_vars["email"].get().strip(),
            "note_entete": self.text_vars["note_entete"].get().strip(),
        }
        self.app.config_data["bulletin_pied_de_page"] = self.footer_text.get("1.0", "end").strip()
        self.app.config_data["entreprise"] = self.text_vars["nom_entreprise"].get().strip() or "Mon Entreprise"

        storage.save(self.app.config_data)
        messagebox.showinfo("Enregistré", "Paramètres de paie mis à jour.")

    def open_data_folder(self):
        import subprocess, sys, os as _os
        path = storage.get_data_dir()
        try:
            if sys.platform.startswith("win"):
                _os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception:
            messagebox.showinfo("Dossier de données", path)


# ==========================================================================
# ONGLET SÉCURITÉ (Administrateur uniquement)
# ==========================================================================

class SecurityTab(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app

        frame = ttk.Frame(self)
        frame.pack(padx=20, pady=20, anchor="nw")

        ttk.Label(frame, text="Mot de passe Utilisateur du mois en cours",
                  font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        period = auth.current_period()
        current_pwd = auth.get_effective_user_password(app.config_data)
        ttk.Label(frame, text=f"Mois : {auth.period_label(period)}").grid(row=1, column=0, sticky="w")
        self.pwd_display = tk.StringVar(value=current_pwd)
        entry = ttk.Entry(frame, textvariable=self.pwd_display, width=20, state="readonly",
                           font=("Consolas", 12, "bold"))
        entry.grid(row=2, column=0, sticky="w", pady=6)
        ttk.Label(frame, text="(généré automatiquement — change chaque 1er du mois)",
                  foreground="#666").grid(row=3, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text="Propre à cette installation : communiquez ce code à l'utilisateur\n"
                               "de cet ordinateur à chaque changement de mois (téléphone, SMS...).",
                  foreground="#666", justify="left").grid(row=4, column=0, columnspan=2, sticky="w", pady=(4, 0))

        ttk.Separator(frame).grid(row=5, column=0, columnspan=2, sticky="ew", pady=16)

        ttk.Label(frame, text="Forcer un mot de passe Utilisateur pour ce mois-ci",
                  font=("Segoe UI", 11, "bold")).grid(row=6, column=0, columnspan=2, sticky="w")
        self.override_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.override_var, width=20).grid(row=7, column=0, sticky="w", pady=6)
        ttk.Button(frame, text="Appliquer", command=self.apply_override).grid(row=7, column=1, padx=8)
        ttk.Button(frame, text="Revenir à la génération automatique",
                   command=self.clear_override).grid(row=8, column=0, columnspan=2, sticky="w", pady=(0, 6))

        ttk.Separator(frame).grid(row=9, column=0, columnspan=2, sticky="ew", pady=16)

        ttk.Label(frame, text="Changer le mot de passe Administrateur",
                  font=("Segoe UI", 11, "bold")).grid(row=10, column=0, columnspan=2, sticky="w")
        self.new_admin_pwd = tk.StringVar()
        ttk.Entry(frame, textvariable=self.new_admin_pwd, width=20, show="•").grid(
            row=11, column=0, sticky="w", pady=6)
        ttk.Button(frame, text="Changer", command=self.change_admin_password).grid(row=11, column=1, padx=8)

        ttk.Separator(frame).grid(row=12, column=0, columnspan=2, sticky="ew", pady=16)
        ttk.Label(frame, text=PAID_SOFTWARE_NOTICE, foreground="#b8860b",
                  font=("Segoe UI", 9, "italic")).grid(row=13, column=0, columnspan=2, sticky="w")

    def apply_override(self):
        pwd = self.override_var.get().strip()
        if not pwd:
            messagebox.showerror("Erreur", "Saisissez un mot de passe.")
            return
        period = auth.current_period()
        self.app.config_data.setdefault("user_password_overrides", {})[period] = pwd
        storage.save(self.app.config_data)
        self.pwd_display.set(pwd)
        messagebox.showinfo("Appliqué", f"Mot de passe Utilisateur forcé pour {auth.period_label(period)}.")

    def clear_override(self):
        period = auth.current_period()
        self.app.config_data.get("user_password_overrides", {}).pop(period, None)
        storage.save(self.app.config_data)
        auto_pwd = auth.get_effective_user_password(self.app.config_data)
        self.pwd_display.set(auto_pwd)
        messagebox.showinfo("Réinitialisé", "Le mot de passe Utilisateur est de nouveau généré automatiquement.")

    def change_admin_password(self):
        pwd = self.new_admin_pwd.get().strip()
        if len(pwd) < 6:
            messagebox.showerror("Erreur", "Le mot de passe doit contenir au moins 6 caractères.")
            return
        salt, digest = auth.hash_password(pwd)
        self.app.config_data["admin_salt"] = salt
        self.app.config_data["admin_hash"] = digest
        storage.save(self.app.config_data)
        self.new_admin_pwd.set("")
        messagebox.showinfo("Changé", "Mot de passe administrateur mis à jour.")


# ==========================================================================

if __name__ == "__main__":
    app = App()
    app.mainloop()
